import openpyxl
from pypdf import PdfReader
import sys
import os


class Check_SWS_Ids:
	
	MAX_CELL_ROW = 150
	SWS_COL      = 3
	
	def __init__(self, Module, PdfFileName, ExcelFileName):
		Pdf_SWS_Ids = self.getSWSIdsFromPdf(Module, PdfFileName)
		#print(Pdf_SWS_Ids)
		Excel_SWS_Ids = self.getSWSIdsFromExcel(ExcelFileName)
		#print(Excel_SWS_Ids)

		print("SWS Ids found   = %d" % len(Pdf_SWS_Ids))
		print("Excel Ids found = %d" % len(Excel_SWS_Ids))
		print("------------------------------------------------------")

		MissingInExcel = self.missingInExcel(Pdf_SWS_Ids, Excel_SWS_Ids)
		if len(MissingInExcel)!=0:
			print("%d present in SWS but missing in Excel" % (len(MissingInExcel))) 
			print(MissingInExcel)
		else:
			print("No ids present in SWS but missing in Excel")
		print("------------------------------------------------------")

		MissingInSWS = self.missingInSWS(Pdf_SWS_Ids, Excel_SWS_Ids)
		if len(MissingInSWS)!=0:
			print("%d present in Excel but missing in SWS" % (len(MissingInSWS))) 
			print(MissingInSWS)
		else:
			print("No ids present in Excel but missing in SWS")
		print("------------------------------------------------------")

	def missingInExcel(self, Pdf_SWS_Ids, Excel_SWS_Ids):
		MissingIds = []
		for PdfFileId in Pdf_SWS_Ids:
			Found = False
			for ExcelFileId in Excel_SWS_Ids:
				if PdfFileId in ExcelFileId:
					Found = True
					break
			
			if Found==False:
				MissingIds.append(PdfFileId)
		
		return(MissingIds)

	def missingInSWS(self, Pdf_SWS_Ids, Excel_SWS_Ids):
		MissingIds = []

		for ExcelFileId in Excel_SWS_Ids:
			Found = False
			for PdfFileId in Pdf_SWS_Ids:
				if PdfFileId in ExcelFileId:
					Found = True
					break
			
			if Found==False:
				MissingIds.append(ExcelFileId)
		
		return(MissingIds)
		
	def getSWSIdsFromPdf(self, Module, PdfFileName):            
		try:
			reader = PdfReader(PdfFileName)
		except:
			print("Unable to open %s" % PdfFileName)
			exit(0)

		number_of_pages = len(reader.pages)
		SWSIds = []
		for page_no in range(0,number_of_pages):
			page  = reader.pages[page_no]
			text  = page.extract_text()
			lines = text.split('\n')

			for line in lines:
				SearchTag = "[SWS_" + Module + "_"
				#print(SearchTag)
				if SearchTag in line:
					Id = line[(line.index('[')+1):line.index(']')]
					#print(Id)
					if self.isIdNotPresent(Id, SWSIds):
						SWSIds.append(Id)                
		
		return(SWSIds)
	
	def isIdNotPresent(self, Id, SWSIds):
		#print(Id)
		for SWSId in SWSIds:
			if Id==SWSId:
				return(False)
		
		return(True)
		
	def getSWSIdsFromExcel(self, ExcelFileName):
		wb = openpyxl.load_workbook(filename=ExcelFileName)
		try:
			wb = openpyxl.load_workbook(filename=ExcelFileName)
		except:
			print("[E] Cannot open %s" % ExcelFileName)
			exit(0)
						
		try:
			SWSWb = wb['SWS']
		except:
			print("[E] No SWS worksheet present")
			exit(0)
			
		SWSIds = []
		TagRowFound = False
		for i in range(1, self.MAX_CELL_ROW):
			SWSId   = SWSWb.cell(row=i, column=self.SWS_COL).value
			#print(SWSId)
			
			if TagRowFound==True:
				if SWSId is not None:
					SWSIds.append(SWSId)
			elif SWSId is not None: 
				if SWSId=="SWS Tag":
					TagRowFound = True
		
		return(SWSIds)
			
if __name__ == "__main__":	        
	#Check_SWS_Ids("Mcu", "D:\\work\\30_info\\AUTOSAR\\00_classic\\4.3.1\\MCAL\\AUTOSAR_SWS_MCUDriver.pdf", "D:\\work\\30_info\\Projects\\AUTOSAR\\LPC2129\\MCAL\\Mcu\\requirements\\LPC2129_Mcu_SRS_SWS.xlsx")
	if len(sys.argv) == 4:
		Check_SWS_Ids(sys.argv[1], sys.argv[2], sys.argv[3])
	else:
		print("[E] Usage: check_sws_ids.py <Module> <SWSPdfFileName> <ExcelFileName>")        