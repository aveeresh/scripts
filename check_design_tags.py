import openpyxl
from pypdf import PdfReader
import sys
import os

class Check_Design_Tagging:
	
	MAX_CELL_ROW = 1000
	SWS_COL      = 3
	IMPL_COL     = 5
	
	def __init__(self, Module, ExcelFileName, PdfFileName):        
		Design_SWS_Ids = self.getSWSIdsFromDesign(Module, PdfFileName)
		Excel_SWS_Ids = self.getSWSIdsFromExcel(ExcelFileName)
		#print(Excel_SWS_Ids)
		
		print("Tags found in Excel  = %d" % len(Excel_SWS_Ids))
		print("Tags found in Design = %d" % len(Design_SWS_Ids))
		
		UntaggedIds = []
		for ExcelFileId in Excel_SWS_Ids:
			Found = False
			for SrcFileId in Design_SWS_Ids:
				if ExcelFileId in SrcFileId:
					Found = True
					break
			
			if Found==False:
				UntaggedIds.append(ExcelFileId)
		
		print("%d untagged Excel Ids found" % len(UntaggedIds)) 
		print(UntaggedIds)
		
	def getSWSIdsFromDesign(self, Module, PdfFileName):            
		try:
			reader = PdfReader(PdfFileName)
		except:
			print("Unable to open %s" % PdfFileName)
			exit(0)

		number_of_pages = len(reader.pages)
		SWSIds = []
		for page_no in range(0,number_of_pages):
			page  = reader.pages[page_no]
			text  = page.extract_text()
			lines = text.split('\n')

			for line in lines:
				SearchTag = "[SWS_" + Module + "_"
				#print(SearchTag)
				if SearchTag in line:
					Id = line[(line.index('[')+1):line.index(']')]
					#print(Id)
					if self.isIdNotPresent(Id, SWSIds):
						SWSIds.append(Id)                
		
		return(SWSIds)

	def isIdNotPresent(self, Id, SWSIds):
		#print(Id)
		for SWSId in SWSIds:
			if Id==SWSId:
				return(False)
		
		return(True)
	
	def getSWSIdsFromExcel(self, ExcelFileName):
		try:
			wb = openpyxl.load_workbook(filename=ExcelFileName)
		except:
			print("[E] Cannot open %s" % ExcelFileName)
			exit(0)
						
		try:
			SWSWb = wb['SWS']
		except:
			print("[E] No SWS worksheet present")
			exit(0)
			
		SWSIds = []
		TagRowFound = False
		for i in range(1, self.MAX_CELL_ROW):
			SWSId   = SWSWb.cell(row=i, column=self.SWS_COL).value
			ImplVal = SWSWb.cell(row=i, column=self.IMPL_COL).value
			#print(SWSId)
			#print(ImplVal)
			
			if TagRowFound==True:
				if SWSId is not None:
					if ImplVal=="YES":
						SWSIds.append(SWSId)
			elif SWSId is not None: 
				if SWSId=="SWS Tag":
					TagRowFound = True
		
		return(SWSIds)
			
if __name__ == "__main__":
	if len(sys.argv) == 4:
		Check_Design_Tagging(sys.argv[1], sys.argv[2], sys.argv[3])
	else:
		print("[E] Usage: check_design_tags.py <Module> <ExcelFile> <PdfDesignFile>")        