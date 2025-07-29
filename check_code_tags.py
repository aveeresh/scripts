import openpyxl
import sys
import os

class Check_Code_Tags:
    
    MAX_CELL_ROW = 150
    SWS_COL      = 3
    IMPL_COL     = 5
    
    def __init__(self, ExcelFileName, DirPath):        
        Src_SWS_Ids = self.getSWSIdsFromSrc(DirPath)
        Excel_SWS_Ids = self.getSWSIdsFromExcel(ExcelFileName)
        #print(Excel_SWS_Ids)
        
        UntaggedIds = []
        for ExcelFileId in Excel_SWS_Ids:
            Found = False
            for SrcFileId in Src_SWS_Ids:
                if ExcelFileId in SrcFileId:
                    Found = True
                    break
            
            if Found==False:
                UntaggedIds.append(ExcelFileId)
        
        print("-----------------------------------------------------")
        print("Total tags found in Excel = %d" % len(Excel_SWS_Ids))
        print("-----------------------------------------------------")
        
        print("%d untagged SWS Ids found" % len(UntaggedIds)) 
        print(UntaggedIds)
        print("-----------------------------------------------------")
        
    def getSWSIdsFromSrc(self, DirPath):
        SrcFilesList = []
        
        for root, dirs, files in os.walk(DirPath):
            for file in files:
                #print(file)
                if file.endswith(".c") or file.endswith(".h"):
                    SrcFilesList.append(os.path.join(root, file))
        
        SWSIds = []
        for File in SrcFilesList:
            with open(File, "r") as FileHandle:
                lines = FileHandle.readlines()
                for line in lines:
                    if "SWS_" in line:
                        #print(line)
                        SWSIds.append(line)
        
        return(SWSIds)
    
    def getSWSIdsFromExcel(self, ExcelFileName):
        try:
            wb = openpyxl.load_workbook(filename=ExcelFileName)
        except:
            print("[E] Cannot open %s" % ExcelFileName)
            exit(0)
                        
        try:
            SWSWb = wb['SWS']
        except:
            print("[E] No SWS worksheet present")
            exit(0)
            
        SWSIds = []
        TagRowFound = False
        for i in range(1, self.MAX_CELL_ROW):
            SWSId   = SWSWb.cell(row=i, column=self.SWS_COL).value
            ImplVal = SWSWb.cell(row=i, column=self.IMPL_COL).value
            #print(SWSId)
            #print(ImplVal)
            
            if TagRowFound==True:
                if SWSId is not None:
                    if ImplVal=="YES":
                        SWSIds.append(SWSId)
            elif SWSId is not None: 
                if SWSId=="SWS Tag":
                    TagRowFound = True
        
        return(SWSIds)
            
if __name__ == "__main__":
	if len(sys.argv) == 3:
		Check_Code_Tags(sys.argv[1], sys.argv[2])
	else:
		print("[E] Usage: check_code_tags.py <ExcelFileName> <SearchDirPath>")        